__all__ = ["BuildLabels"]
from copy import deepcopy
from io import BytesIO
from logging import getLogger
import math
import re

from babel.dates import format_datetime
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
import pandas as pd

from api.models import BuildLabelsResponse


logger = getLogger(__name__)
logger.setLevel("DEBUG")


class BuildLabels:
    """
    Class to build labels for flows.
    """

    def __init__(self, columna_n_duplicados: str, grid_template: bytes, date_format: str):
        self.columna_n_duplicados = columna_n_duplicados
        self.grid_template = BeautifulSoup(grid_template, "html.parser")
        self.date_format = date_format
        self.n_etiquetas_por_hoja = 4
        self.errors = []

    def run(self, data, template: str) -> list[str]:
        """
        Build labels for the flow.
        """
        
        labels_template = BeautifulSoup(template, "html.parser")
        grid_template = self._extend_styles(self.grid_template, labels_template)
        data = load_workbook(filename=BytesIO(data), rich_text=True)
        data = self._extract_format(data)
        data = self._validate_data(data)
        data = self._enumerate_labels(data)
        labels = self._fill_template(data, labels_template)
        grid = self._fill_grid(labels, grid_template)

        return BuildLabelsResponse(labels=grid.prettify(), errors=list(set(self.errors)))

    def _validate_data(self, data):
        logger.info(f"Total de datos en el archivo: {data.shape}")
        if data.empty:
            return_msg = "El archivo está vacío."
            raise ValueError(return_msg)
        
        data["_label_id"] = range(len(data))

        data = self._filter_labels_number(data)
        if data.empty:
            logger.warning("No hay ejemplares con etiquetas.")
            return_msg = f"""
                No hay ejemplares con etiquetas. 
                Se requiere al menos un duplicado para crear etiquetas. 
                Asegurese que exista la columna '{self.columna_n_duplicados}' y que tenga al menos un valor mayor a cero.
            """
            raise ValueError(return_msg)
        data.columns = data.columns.str.strip()
        return data

    def _filter_labels_number(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Filter specimens with at least one duplicate.
        """
        column_name = self.columna_n_duplicados
        if column_name not in data.columns:
            raise ValueError(f"La columna {column_name} no se encuentra en el archivo")
        data = data.loc[data[column_name] > 0]
        logger.info(f"Creando etiquetas para {len(data)} ejemplares.")
        return data

    def _enumerate_labels(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Enumerate total labels.
        """
        data = data.reindex(data.index.repeat(data[self.columna_n_duplicados]))
        logger.info(f"Total de etiquetas por crear: {len(data)}")
        return data

    def _extend_styles(self, grid: BeautifulSoup, labels: str) -> BeautifulSoup:
        """
        Extend styles in the grid.
        """
        grid.find("style").extend(labels.find("style").contents)
        return grid

    def _extract_format(self, workbook: Workbook) -> pd.DataFrame:
        # The 'rich_text=True' parameter is required otherwise the cells are
        sheet = workbook.active
        for row in sheet.iter_rows():
            for cell in row:
            # Check if the entire cell is italicized
                if cell.font.italic:
                    cell.value = f"<em>{cell.value}</em>"
            # cell.value will either be CellRichText or str, with CellRichText having more formatting that needs to be checked.
                if isinstance(cell.value, CellRichText):
                    for text_block in cell.value:
                    # Ensure it's a text block not a plain string, and that it is in fact italicized
                        if isinstance(text_block, TextBlock) and text_block.font.italic:
                            text_block.text = f"<em>{text_block.text}</em>"

        df = pd.DataFrame(list(sheet.values))
        df.columns = df.iloc[0]  
        df = df[1:]  
        new_col_names = []
        for col_name in df.columns:
            if isinstance(col_name, TextBlock):
                col_name = col_name.text
            elif isinstance(col_name, CellRichText):
                col_name = str(col_name)
            new_col_names.append(col_name)

        df.columns = new_col_names
        df.columns = df.columns.str.replace("<em>", "")  
        df.columns = df.columns.str.replace("</em>", "")
        workbook.close()
        return df

    def _fill_template(self, data: pd.DataFrame, template: str) -> list[str]:
        """
        Replace templates with data.
        """
        logger.info("Creando templates de etiquetas.")
        data["_template"] = data.apply(
            self._replace_templates, axis=1, template=template
        )
        return data["_template"].tolist()

    def _replace_templates(self, data: pd.Series, template: BeautifulSoup) -> str:
        """
        Fill the template with data.
        """
        html_str = deepcopy(template).prettify()
        for column in data.index:
            if column is None:
                continue
            replace_str = "#{" + column + "}#"
            value = data[column]

            if isinstance(value, datetime):
                value = format_datetime(value, self.date_format, locale="es")
            if value is None:
                value = ""
            
            html_str = html_str.replace(replace_str, str(value).strip())
        
        pattern = r"#\{.*?\}#"
        matches = re.findall(pattern, html_str)
        if matches:
            error_msg = f"Valores faltantes para el ejemplar {data._label_id}: {matches}"
            self.errors.append(error_msg)
            logger.warning(error_msg)
            for match in matches:
                html_str = html_str.replace(match, "")
        return html_str

    def _fill_grid(self, labels: list[str], grid_template: str) -> BeautifulSoup:
        """
        Fill the grid with labels.
        """
        logger.info("Creando grid de etiquetas.")
        grid = deepcopy(grid_template)
        n_pages = math.ceil(len(labels) / self.n_etiquetas_por_hoja)
        logger.info(f"Total de páginas: {n_pages}")
        labels = [BeautifulSoup(label, "html.parser") for label in labels]
        grid.find("div", {"class": "label-grid"}).clear()
        grid.find("div", {"class": "label-grid"}).extend(
            [label.body.find("div", {"class": "label-body"}) for label in labels]
        )
        grid.find("title").string = "Etiquetas"
        return grid

    def _save_labels(self, grid: BeautifulSoup, path: str) -> None:
        with open(self.archivo_etiquetas, "w", encoding="utf-8") as f:
            f.write(grid.prettify())
        logger.info(f"Etiquetas guardadas en {path}")
